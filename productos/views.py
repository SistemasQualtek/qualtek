# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.core.urlresolvers import reverse_lazy
from django.conf import settings
from django.views.generic import ListView, View
from django.views.generic.detail import DetailView
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.shortcuts import render, get_object_or_404, render_to_response
from .models import Producto, Proveedor, Log
from .forms import ProductoForm, ProveedorForm, LogForm
from clientes.models import Venta, Prod_Cli, Cliente
from django.views.generic.base import TemplateView
import pprint
import datetime
from django.views.generic.edit import (
    CreateView,
    UpdateView,
    DeleteView
)
import os
from reportlab.pdfgen import canvas
from django.views.generic import ListView
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table
from io import BytesIO
from reportlab.lib.colors import PCMYKColor
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend
from django.db.models import Q, Sum
import time
from openpyxl import Workbook


def ProductoList(request):
    count = Producto.objects.count()
    productos = Producto.objects.order_by('proveedor','medida')
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()
            producto.save()
            return HttpResponseRedirect('/Lista/Productos/')
    else:
        form = ProductoForm()
    template = loader.get_template('productos/producto_list.html')
    context = {
        'productos':productos,
        'form':form,
        'count':count,
    }
    # print (productos)
    return render(request, 'productos/producto_list.html', context)

def ProductoDetail(request,pk):
    producto = get_object_or_404(Producto, pk=pk)
    template = loader.get_template('productos/producto_detail.html')
    forma = LogForm(request.POST, request.FILES)
    if forma.is_valid():
        cantidad = forma.cleaned_data['cantidad']
        venta = Log()
        producto.existencia = producto.existencia - cantidad
        venta.producto = producto
        venta.cantidad = cantidad
        producto.save()
        venta.save()
    context = {
    'forma': forma,
    'producto': producto
    }
    if request.user.is_authenticated():
        return HttpResponse(template.render(context, request))
    else:
        return HttpResponseRedirect(reverse('productos/producto_detail.html', args=(producto.id,)))
# Inicio de Salida Manual
def Salida(request,pk):
    producto = get_object_or_404(Producto, pk=pk)
    nombre = 'Salida'
    template = loader.get_template('productos/producto_detail.html')
    forma = LogForm(request.POST, request.FILES)
    if forma.is_valid():
        cantidad = forma.cleaned_data['cantidad']
        culpable = forma.cleaned_data['culpable']
        ilicito = forma.cleaned_data['ilicito']
        venta = Log()
        producto.existencia = producto.existencia - cantidad
        venta.producto = producto
        venta.cantidad = cantidad
        venta.culpable = culpable
        venta.ilicito = ilicito
        producto.save()
        venta.save()
    context = {
    'forma': forma,
    'producto': producto,
    'nombre': nombre
    }
    if request.user.is_authenticated():
        return HttpResponse(template.render(context, request))
    else:
        return HttpResponseRedirect(reverse('productos/producto_detail.html', args=(producto.id,)))
# Fin de Salida Manual

# Inicio de Entrada Manual
def Entrada(request,pk):
    producto = get_object_or_404(Producto, pk=pk)
    nombre = 'Entrada'
    template = loader.get_template('productos/producto_detail.html')
    forma = LogForm(request.POST, request.FILES)
    if forma.is_valid():
        cantidad = forma.cleaned_data['cantidad']
        culpable = forma.cleaned_data['culpable']
        ilicito = forma.cleaned_data['ilicito']
        venta = Log()
        producto.existencia = producto.existencia + cantidad
        venta.producto = producto
        venta.cantidad = cantidad
        venta.culpable = culpable
        venta.ilicito = ilicito
        producto.save()
        venta.save()
    context = {
    'forma': forma,
    'producto': producto,
    'nombre': nombre
    }
    if request.user.is_authenticated():
        return HttpResponse(template.render(context, request))
    else:
        return HttpResponseRedirect(reverse('productos/producto_detail.html', args=(producto.id,)))
# Fin de Entrada Manual

class ProductoUpdate(UpdateView):
    model = Producto
    success_url = reverse_lazy('almacen:producto_list')
    fields = ['codigo', 'descripcion','medida','unidad', 'proveedor','ubicacion','cantidad_caja', 'cantidad_rb', 'existencia', 'costo','precio', 'release']
class ProductoDelete(DeleteView):
    model = Producto
    success_url = reverse_lazy('almacen:producto_list')
    fields = ['codigo', 'descripcion', 'proveedor', 'existencia', 'costo', 'precio']

def principal(request):
    playera = Producto.objects.order_by('id')
    template = loader.get_template('home_page.html')
    context = {
        'playera': playera
    }
    return HttpResponse(template.render(context, request))

def inicio(request):
    playera = Producto.objects.order_by('id')
    template = loader.get_template('inicio.html')
    context = {
        'playera': playera
    }
    return HttpResponse(template.render(context, request))

def producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()
            producto.save()
            return HttpResponseRedirect('/productos/Lista/Productos/')
    else:
        form = ProductoForm()
    template = loader.get_template('productos/new_producto.html')
    context = {
        'form': form
    }
    if request.user.is_authenticated():
        return HttpResponse(template.render(context, request))
    else:
        return HttpResponseRedirect('productos:producto_list')

def categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST, request.FILES)
        if form.is_valid():
            playera = form.save()
            playera.save()
            return HttpResponseRedirect('/productos/Lista/Productos/')
    else:
        form = CategoriaForm()
    template = loader.get_template('productos/new_categoria.html')
    context = {
        'form': form
    }
    if request.user.is_authenticated():
        return HttpResponse(template.render(context, request))
    else:
        return HttpResponseRedirect('productos:productos_list')

def Lista_Log(request):
    log = Log.objects.filter(fecha= time.strftime("%Y-%m-%d"))
    template = loader.get_template('productos/venta_list.html')
    context = {
        'log': log
    }
    return HttpResponse(template.render(context, request))


def ReleaseList(request):
    count = Producto.objects.count()
    productos = Producto.objects.order_by('id')
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()
            producto.save()
            return HttpResponseRedirect('almacen:release')
    else:
        form = ProductoForm()
    template = loader.get_template('productos/release.html')
    context = {
        'productos':productos,
        'form':form,
        'count':count

    }
    # print (productos)
    return render(request, 'productos/release.html', context)

####################################### ahuevo ################################

def pdftwm(request):
    # print "Genero el PDF Ventas del Dia"
    date = request.GET.get('date','2017-05-04')
    medida = request.GET.get('medida','2017-05-04')
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Tubo_W_con_Medida_de_"+medida+".pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    courses = []
    styles = getSampleStyleSheet()
    header = Paragraph("Tubo W de "+medida, styles['Title'])
    courses.append(header)
    headings = ('Codigo', 'Descripcion', 'Medida', 'Existencia')
    allcourses = [(p.codigo, p.descripcion, p.medida, p.existencia) for p in Producto.objects.filter(medida = medida).order_by('descripcion')]
    t = Table([headings] + allcourses)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (3, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white)
        ]
    ))

    courses.append(t)
    doc.build(courses)
    response.write(buff.getvalue())
    buff.close()
    return response
####################################### ahuevo ################################

####################################### Proveedores Pedidos ########################
def PedidosList(request):

    template = loader.get_template('productos/pedidos_list.html')
    context = {
    
    }
    # print (productos)
    return render(request, 'productos/pedidos_list.html', context)


################################## Proveedor #######################################
def ProveedorList(request):
    count = Proveedor.objects.count()
    proveedor = Proveedor.objects.order_by('id')
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save()
            producto.save()
            return HttpResponseRedirect('/Lista/Proveedores/')
    else:
        form = ProveedorForm()
    template = loader.get_template('proveedores/proveedor_list.html')
    context = {
        'proveedor':proveedor,
        'form':form,
        'count':count

    }
    # print (productos)
    return render(request, 'proveedores/proveedor_list.html', context)

from django.db import connection

def ProveedorDetail(request,pk):
    elculodealberto = get_object_or_404(Proveedor, pk=pk)
    template = loader.get_template('proveedores/proveedor_detail.html')
    cursor = connection.cursor()
    cursor.execute('''SELECT * FROM pronostico_agosto_a3meses''')
    for i in [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14]:
        row = cursor.fetchone()
        print i
        for i in [0,1,2,3,4,5,6,7,8,9]:
            print row[i]
    # print row[1]


    context = {
    'row':row,
    'elculodealberto':elculodealberto
    }
    if request.user.is_authenticated():
        return HttpResponse(template.render(context, request))
    else:
        return HttpResponseRedirect(reverse('proveedores/proveedor_detail.html', args=(elculodealberto.id,)))

################################################## EXCEL Reporte ##########################################################
class Reportar_Pedidos(TemplateView):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        cursora = connection.cursor()
        cursora.execute('''SELECT COUNT(*) FROM pronostico_a3meses''')
        rowa = cursora.fetchone()
        cursor = connection.cursor()
        j = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,34,35,36,37,38,39,40,41,42,43,44,45,46]
        cursor.execute('''SELECT * FROM pronostico_a3meses''')
        ws['A1'] = 'Proveedor'
        ws['B1'] = 'Codigo'
        ws['C1'] = 'Producto'
        ws['D1'] = 'Unidad'
        ws['E1'] = 'Existencia'
        ws['F1'] = 'Vendido en Agosto'
        ws['G1'] = 'Stock post agosto'
        ws['H1'] = 'pronostico a 3 meses'
        ws['I1'] = 'Stock post 3 meses'
        cont=2
        for i in j:
            row = cursor.fetchone()
            for con in rowa:
                ws.cell(row=cont,column=1).value = row[0]
                ws.cell(row=cont,column=2).value = row[1]
                ws.cell(row=cont,column=3).value = row[2]
                ws.cell(row=cont,column=4).value = row[3]
                ws.cell(row=cont,column=5).value = row[4]
                ws.cell(row=cont,column=6).value = row[5]
                ws.cell(row=cont,column=7).value = row[6]
                ws.cell(row=cont,column=8).value = row[7]
                ws.cell(row=cont,column=9).value = row[8]
                cont = cont + 1
        nombre_archivo ="pedidos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
