# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.core.urlresolvers import reverse_lazy
from django.conf import settings
from django.views.generic import ListView, View
from django.views.generic.detail import DetailView
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.shortcuts import render, get_object_or_404, render_to_response
from productos.models import Producto, Proveedor, Log
from productos.forms import ProductoForm, ProveedorForm, LogForm
from django.views.generic.base import TemplateView
import pprint
from django.views.generic.edit import (
    CreateView,
    UpdateView,
    DeleteView
)
import os
from reportlab.pdfgen import canvas
from django.views.generic import ListView
from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table
from io import BytesIO
from reportlab.lib.colors import PCMYKColor
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend
from django.db.models import Q, Sum
import time
from openpyxl import Workbook

###############################################REPORTES######################################
###############################################QUALTEK######################################

###############################################GENERAL######################################
def pdfgen(request):
    # print "Genero el PDF"
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Stock.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    courses = []
    styles = getSampleStyleSheet()
    header = Paragraph("Reporte General de Almacen.", styles['Title'])
    header.hAlign = 'CENTER'
    courses.append(header)
    headings = ('Código', 'Descripción','Medida','Unidad', 'Existencia', 'Proveedor')
    allcourses = [(p.codigo, p.descripcion, p.medida, p.unidad, p.existencia, p.proveedor) for p in Producto.objects.all().order_by('proveedor', 'medida')]
    # print allcourses
    t = Table([headings] + allcourses)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (5, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('ALIGN',(0,0),(3,0),'CENTER'),
        ]
    ))

    courses.append(t)
    doc.build(courses)
    response.write(buff.getvalue())
    buff.close()
    return response
###############################################GENERAL######################################

####################################### Reporte Cinchos ####################################
def pdfcin(request):
    # print "Genero el PDF"
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "cinchos.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    courses = []
    styles = getSampleStyleSheet()
    header = Paragraph("Reporte de Cinchos.", styles['Title'])
    header.hAlign = 'CENTER'
    courses.append(header)
    headings = ('Código', 'Descripción','Medida','Unidad', 'Existencia', 'Proveedor')
    allcourses = [(p.codigo, p.descripcion, p.medida, p.unidad, p.existencia, p.proveedor) for p in Producto.objects.all().filter(proveedor='Qualtek').order_by('medida')]
    # print allcourses
    t = Table([headings] + allcourses)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (5, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('ALIGN',(0,0),(3,0),'CENTER'),
        ]
    ))

    courses.append(t)
    doc.build(courses)
    response.write(buff.getvalue())
    buff.close()
    return response
#################################### Fin Reporte Cinchos ##############################################
####################################### Reporte Tubo Qualtek ####################################
def pdftq(request):
    # print "Genero el PDF"
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "tuboqualtek.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    courses = []
    styles = getSampleStyleSheet()
    header = Paragraph("Qualtek México S.A. de C.V.", styles['Title'])
    qualtek = Paragraph("Reporte Tubo Qualtek.", styles['Heading2'])
    header.hAlign = 'CENTER'
    courses.append(header)
    courses.append(qualtek)
    headings = ('Código', 'Descripción','Medida','Unidad', 'Existencia', 'Proveedor')
    allcourses = [(p.codigo, p.descripcion, p.medida, p.unidad, p.existencia, p.proveedor) for p in Producto.objects.all().filter(proveedor='Tubo Qualtek').order_by('medida')]
    # print allcourses
    t = Table([headings] + allcourses)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (5, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('ALIGN',(0,0),(3,0),'CENTER'),
        ]
    ))

    courses.append(t)
    doc.build(courses)
    response.write(buff.getvalue())
    buff.close()
    return response
#################################### Fin Reporte Tubo Qualtek##############################################

#######################################ReporteTubo W###########################
def pdfW(request):
    # print "Genero el PDF"
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "tubow.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    courses = []
    styles = getSampleStyleSheet()
    header = Paragraph("Qualtek México S.A. de C.V.", styles['Title'])
    qualtek = Paragraph("Reporte Tubo W.", styles['Heading2'])
    header.hAlign = 'CENTER'
    courses.append(header)
    courses.append(qualtek)
    headings = ('Código', 'Descripción','Medida','Unidad', 'Existencia', 'Proveedor')
    allcourses = [(p.codigo, p.descripcion, p.medida, p.unidad, p.existencia, p.proveedor) for p in Producto.objects.all().filter(proveedor='Tubo W').order_by('medida')]
    # print allcourses
    t = Table([headings] + allcourses)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (5, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white),
            ('ALIGN',(0,0),(3,0),'CENTER'),
        ]
    ))

    courses.append(t)
    doc.build(courses)
    response.write(buff.getvalue())
    buff.close()
    return response
#################################### Fin Reporte Tubo Qualtek##############################################



def pdfrel(request):
    # print "Genero el PDF"
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Release.pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    # response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    courses = []
    styles = getSampleStyleSheet()
    header = Paragraph("Release", styles['Title'])
    header.Align = 'CENTER'
    courses.append(header)
    headings = ('Código', 'Descripción', 'Existencia', 'Proveedor')
    allcourses = [(p.codigo, p.descripcion, p.existencia, p.proveedor) for p in Producto.objects.all().filter(release=True).order_by('id')]
    # print allcourses
    t = Table([headings] + allcourses)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (3, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white)
        ]
    ))

    courses.append(t)
    doc.build(courses)
    response.write(buff.getvalue())
    buff.close()
    return response


def pdfdia(request):
    # print "Genero el PDF Ventas del Dia"
    date = request.GET.get('date','2017-05-04')
    response = HttpResponse(content_type='application/pdf')
    pdf_name = "Reporte del Día "+date+".pdf"  # llamado clientes
    # la linea 26 es por si deseas descargar el pdf a tu computadora
    response['Content-Disposition'] = 'attachment; filename=%s' % pdf_name
    buff = BytesIO()
    doc = SimpleDocTemplate(buff,
                            pagesize=letter,
                            rightMargin=40,
                            leftMargin=40,
                            topMargin=60,
                            bottomMargin=18,
                            )
    courses = []
    styles = getSampleStyleSheet()
    header = Paragraph("Reporte de E/S del "+date, styles['Title'])
    courses.append(header)
    headings = ('Producto', 'Cantidad', 'Ilicito', 'Culpable')
    allcourses = [(p.producto, p.cantidad, p.ilicito, p.culpable) for p in Log.objects.filter(fecha = date)]
    totalfinal = 0
    saldos = 0
    for p in Log.objects.filter(fecha = date):
        saldos = saldos + p.cantidad
        totalfinal = totalfinal + p.cantidad
        # print (saldos)
    total = ('','Total', saldos, totalfinal)
    # print allcourses

    allcourses.append(total)
    t = Table([headings] + allcourses)
    t.setStyle(TableStyle(
        [
            ('GRID', (0, 0), (3, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.white)
        ]
    ))

    courses.append(t)
    doc.build(courses)
    response.write(buff.getvalue())
    buff.close()
    return response


################################################## EXCEL Reporte ##########################################################
class ReporteQualtekExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        conta = Producto.objects.all().order_by('proveedor','medida')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Código'


        ws['B1'] = 'Descripcion'
        ws['C1'] = 'Unidad'
        ws['D1'] = 'Medida'
        ws['E1'] = 'Existencia'
        ws['F1'] = 'Proveedor'
        ws['G1'] = 'Cantidad x Caja'
        ws['H1'] = 'Cantidad x Rollo/Bolsa'
        ws['I1'] = 'Ubicacion'
        ws['J1'] = 'Costo'
        ws['K1'] = 'Precio'
        ws['L1'] = 'Release'
        cont=2

        for con in conta:
            ws.cell(row=cont,column=1).value = con.codigo
            ws.cell(row=cont,column=2).value = con.descripcion
            ws.cell(row=cont,column=3).value = con.unidad
            ws.cell(row=cont,column=4).value = con.medida
            ws.cell(row=cont,column=5).value = con.existencia
            ws.cell(row=cont,column=6).value = con.proveedor
            ws.cell(row=cont,column=7).value = con.cantidad_caja
            ws.cell(row=cont,column=8).value = con.cantidad_rb
            ws.cell(row=cont,column=9).value = con.ubicacion
            ws.cell(row=cont,column=10).value = con.costo
            ws.cell(row=cont,column=11).value = con.precio
            ws.cell(row=cont,column=12).value = con.release
            cont = cont + 1

        nombre_archivo ="AlmacenExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
################################################## EXCEL Reporte ##########################################################
